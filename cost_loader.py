import os

try:
    from openpyxl import load_workbook
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False


def load_material_costs(file_path='data/material_costs_hyd.xlsx'):
    """Load material costs from Excel file.
    
    Args:
        file_path: Path to the Excel file (relative to project root)
    
    Returns:
        Dictionary with material costs in format:
        {
            'cement_per_bag': float,
            'sand_per_m3': float,
            'aggregate_per_m3': float,
            'steel_per_kg': float,
            'brick_per_100': float,
            'tile_per_sqm': float,
            'paint_per_liter': float,
            'plaster_per_m3': float
        }
    """
    
    # Default rates (fallback if file not found)
    default_rates = {
        'cement_per_bag': 420.0,
        'sand_per_m3': 1200.0,
        'aggregate_per_m3': 900.0,
        'steel_per_kg': 65.0,
        'brick_per_100': 350.0,
        'tile_per_sqm': 300.0,
        'paint_per_liter': 500.0,
        'plaster_per_m3': 2000.0
    }
    
    # Check if openpyxl is available
    if not HAVE_OPENPYXL:
        print("openpyxl not installed - using default rates")
        return default_rates

    # Check if file exists
    if not os.path.exists(file_path):
        print(f"Warning: Material costs file not found at {file_path}. Using default rates.")
        return default_rates

    try:
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Try to get the first sheet (or a sheet named 'Materials' or 'Costs')
        sheet_names = wb.sheetnames
        ws = None
        
        for sheet_name in ['Materials', 'Costs', 'Material Costs', 'Rates']:
            if sheet_name in sheet_names:
                ws = wb[sheet_name]
                break
        
        # If no specific sheet found, use the first sheet
        if ws is None:
            ws = wb[sheet_names[0]]
        
        rates = dict(default_rates)
        
        # Parse the sheet
        # Expected format: 
        # Row 1: Headers (Material, Unit, Cost/Price)
        # Subsequent rows: Material data
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:  # Skip empty rows
                continue
            
            material_name = str(row[0]).strip().lower() if row[0] else ''
            unit = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ''
            try:
                cost = float(row[2]) if len(row) > 2 and row[2] else None
            except (ValueError, TypeError):
                cost = None
            
            if not material_name or cost is None:
                continue
            
            # Map material names to our rate keys
            if 'cement' in material_name:
                rates['cement_per_bag'] = cost
            elif 'sand' in material_name:
                rates['sand_per_m3'] = cost
            elif 'aggregate' in material_name or 'gravel' in material_name or 'chips' in material_name:
                rates['aggregate_per_m3'] = cost
            elif 'steel' in material_name or 'rebar' in material_name or 'reinforcement' in material_name:
                rates['steel_per_kg'] = cost
            elif 'brick' in material_name:
                rates['brick_per_100'] = cost
            elif 'tile' in material_name or 'flooring' in material_name:
                rates['tile_per_sqm'] = cost
            elif 'paint' in material_name or 'coating' in material_name:
                rates['paint_per_liter'] = cost
            elif 'plaster' in material_name or 'finish' in material_name:
                rates['plaster_per_m3'] = cost
        
        print(f"Successfully loaded material costs from {file_path}")
        print(f"Loaded rates: {rates}")
        
        wb.close()
        return rates
        
    except Exception as e:
        print(f"Error loading material costs from {file_path}: {str(e)}")
        print("Using default rates as fallback.")
        return default_rates


def get_rates_by_plan(base_rates, plan_type='standard'):
    """Adjust rates based on plan type (economy, standard, premium).
    
    Args:
        base_rates: Base rates loaded from Excel
        plan_type: 'economy', 'standard', or 'premium'
    
    Returns:
        Adjusted rates dictionary
    """
    adjusted = dict(base_rates)
    
    if plan_type == 'economy':
        # Economy plan: reduce rates by 10%
        for key in adjusted:
            adjusted[key] = adjusted[key] * 0.90
    
    elif plan_type == 'premium':
        # Premium plan: increase rates by 15%
        for key in adjusted:
            adjusted[key] = adjusted[key] * 1.15
    
    # Standard plan: no adjustment
    
    return adjusted
